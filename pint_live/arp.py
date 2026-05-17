"""ARP list support — load an IP/MAC/Hostname table from an Excel workbook
and resolve switch-port MACs to their IPs and hostnames.

Excel format expected:
  • First row contains headers
  • Columns named (case-insensitive, flexible): IP / IP Address, MAC / MAC Address,
    Hostname (optional)
  • Any extra columns are ignored
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


# ── MAC normalization ──────────────────────────────────────────────────────

_HEX_RE = re.compile(r"[0-9a-fA-F]")


def normalise_mac(mac: str) -> str:
    """Reduce any MAC representation to bare lowercase hex (12 chars)
    so dotted, colon-, and dash-separated forms compare equal.
    Returns "" if the input doesn't contain exactly 12 hex digits."""
    if not mac:
        return ""
    hex_only = "".join(_HEX_RE.findall(mac)).lower()
    return hex_only if len(hex_only) == 12 else ""


# ── Header matching ────────────────────────────────────────────────────────

_IP_HEADERS       = {"ip", "ip address", "ipaddress", "address"}
_MAC_HEADERS      = {"mac", "mac address", "macaddress", "hardware address"}
_HOSTNAME_HEADERS = {"hostname", "host", "name", "device", "device name"}


def _classify_header(value) -> str | None:
    if value is None:
        return None
    key = str(value).strip().lower()
    if key in _IP_HEADERS:
        return "ip"
    if key in _MAC_HEADERS:
        return "mac"
    if key in _HOSTNAME_HEADERS:
        return "hostname"
    return None


# ── ARP table model ────────────────────────────────────────────────────────

@dataclass
class ArpEntry:
    ip: str
    mac: str          # normalised (lowercase hex, no separators)
    hostname: str = ""


@dataclass
class ArpTable:
    source_path: Path | None = None
    entries: list[ArpEntry] = field(default_factory=list)
    _by_mac: dict[str, ArpEntry] = field(default_factory=dict, repr=False)

    def __post_init__(self) -> None:
        # Last write wins if the file has duplicate MACs.
        for entry in self.entries:
            if entry.mac:
                self._by_mac[entry.mac] = entry

    def __len__(self) -> int:
        return len(self.entries)

    def lookup(self, mac: str) -> ArpEntry | None:
        return self._by_mac.get(normalise_mac(mac))

    def resolve_ips(self, macs: list[str]) -> list[str]:
        """Return one IP per input MAC, in the same order. Unmatched MACs
        produce empty strings, which are kept so position is preserved."""
        out = []
        for m in macs:
            entry = self.lookup(m)
            out.append(entry.ip if entry else "")
        return out

    def resolve_hostnames(self, macs: list[str]) -> list[str]:
        out = []
        for m in macs:
            entry = self.lookup(m)
            out.append(entry.hostname if entry else "")
        return out


# ── Loader ─────────────────────────────────────────────────────────────────

class ArpLoadError(Exception):
    """Raised when the Excel file can't be parsed as an ARP list."""


def load_arp_xlsx(path: Path) -> ArpTable:
    """Load an ARP list from an .xlsx workbook. Reads the active sheet.

    Detects which columns hold IP, MAC, and Hostname from the header row.
    Raises ArpLoadError if the required IP and MAC columns can't be found."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ArpLoadError(f"Could not open workbook: {exc}") from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        raise ArpLoadError("Workbook is empty.")

    col_map: dict[str, int] = {}
    for idx, value in enumerate(header):
        kind = _classify_header(value)
        if kind and kind not in col_map:
            col_map[kind] = idx

    if "ip" not in col_map or "mac" not in col_map:
        raise ArpLoadError(
            "ARP file must have IP and MAC columns. "
            f"Found headers: {[h for h in header if h is not None]}"
        )

    ip_idx       = col_map["ip"]
    mac_idx      = col_map["mac"]
    hostname_idx = col_map.get("hostname")

    entries: list[ArpEntry] = []
    for row in rows:
        if not row:
            continue
        ip  = _cell_str(row, ip_idx)
        mac = normalise_mac(_cell_str(row, mac_idx))
        if not ip or not mac:
            continue
        hostname = _cell_str(row, hostname_idx) if hostname_idx is not None else ""
        entries.append(ArpEntry(ip=ip, mac=mac, hostname=hostname))

    wb.close()
    return ArpTable(source_path=Path(path), entries=entries)


def _cell_str(row, idx: int) -> str:
    if idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    return str(value).strip()
