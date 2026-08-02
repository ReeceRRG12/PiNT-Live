"""Excel workbook exporter using openpyxl."""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pint_live.arp    import ArpTable
from pint_live.models import ParsedSwitchData, InterfaceEntry


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

COLOUR_UP        = "C6EFCE"   # green fill
COLOUR_DOWN      = "FFC7CE"   # red fill
COLOUR_DISABLED  = "D9D9D9"   # grey fill
COLOUR_HEADER    = "1F4E79"   # dark blue header
COLOUR_HEADER_FG = "FFFFFF"
COLOUR_WARNING    = "FFF2CC"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_style(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=COLOUR_HEADER_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=COLOUR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def _link_fill(link: str) -> PatternFill:
    key = link.lower()
    if key == "up":
        return PatternFill("solid", fgColor=COLOUR_UP)
    if key == "disabled":
        return PatternFill("solid", fgColor=COLOUR_DISABLED)
    return PatternFill("solid", fgColor=COLOUR_DOWN)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_summary_link(ws) -> None:
    """Add a consistent internal backlink in the top-left cell."""
    cell = ws["A1"]
    cell.value = "← Back to Summary"
    cell.hyperlink = "#'Summary'!A1"
    cell.style = "Hyperlink"
    cell.font = Font(color="0563C1", underline="single", bold=True, size=10)


def _add_sheet_link(cell, sheet_name: str | None) -> None:
    """Populate a Summary navigation cell when the target sheet exists."""
    if not sheet_name:
        cell.value = "—"
        return
    cell.value = "Open"
    escaped_name = sheet_name.replace("'", "''")
    cell.hyperlink = f"#'{escaped_name}'!A1"
    cell.style = "Hyperlink"


# ---------------------------------------------------------------------------
# Per-switch sheet
# ---------------------------------------------------------------------------

_INTF_HEADERS_BASE = ["Interface", "Link", "State", "Duplex", "Speed", "Untagged VLAN", "Tagged VLANs", "MAC (from table)", "Neighbour Protocol", "Neighbour Device", "Neighbour IP", "Remote Port", "Neighbour Platform", "Description"]
_INTF_WIDTHS_BASE  = [14,          10,     12,      10,       10,      20,              32,              22,                 18,                   28,                 18,             20,            34,                   30]
_INTF_HEADERS_ARP  = ["IP (from ARP)", "Hostname (from ARP)"]
_INTF_WIDTHS_ARP   = [22,              26]


def _write_interfaces_sheet(
    wb: Workbook,
    data: ParsedSwitchData,
    arp_table: ArpTable | None = None,
) -> str:
    tab_name = (data.hostname or data.host)[:31]  # Excel tab name limit
    ws = wb.create_sheet(title=tab_name)

    if arp_table is not None:
        # Insert IP / Hostname directly after the MAC column.
        mac_idx  = _INTF_HEADERS_BASE.index("MAC (from table)")
        headers  = _INTF_HEADERS_BASE[:mac_idx + 1] + _INTF_HEADERS_ARP + _INTF_HEADERS_BASE[mac_idx + 1:]
        widths   = _INTF_WIDTHS_BASE[:mac_idx + 1]  + _INTF_WIDTHS_ARP  + _INTF_WIDTHS_BASE[mac_idx + 1:]
    else:
        headers, widths = _INTF_HEADERS_BASE, _INTF_WIDTHS_BASE
    num_cols = len(headers)

    # --- Summary backlink + title row ---
    _add_summary_link(ws)
    col_span = f"A2:{get_column_letter(num_cols)}2"
    ws.merge_cells(col_span)
    title_cell = ws["A2"]
    title_cell.value = f"{data.model or 'Ruckus ICX'}  |  {data.hostname or data.host}  |  {data.host}  |  Polled: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(bold=True, size=11, color=COLOUR_HEADER)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.append([])  # blank row

    # --- Headers ---
    ws.append(headers)
    for col, header in enumerate(headers, start=1):
        _header_style(ws.cell(row=4, column=col), header)
    ws.row_dimensions[4].height = 18

    # Build a quick lookup: port -> MACs from mac table
    port_macs: dict[str, list[str]] = {}
    for entry in data.mac_table:
        port_macs.setdefault(entry.port, []).append(entry.mac)
    port_neighbors = {}
    for neighbor in data.neighbors:
        port_neighbors.setdefault(neighbor.local_port, []).append(neighbor)

    # --- Data rows ---
    for intf in data.interfaces:
        macs_list = port_macs.get(intf.port, [])
        macs = ", ".join(macs_list)
        neighbors = port_neighbors.get(intf.port, [])
        row = [
            intf.port,
            intf.link,
            intf.state,
            intf.duplex,
            intf.speed,
            intf.untagged_vlan,
            intf.tagged_vlans,
            macs,
        ]
        if arp_table is not None:
            ips       = arp_table.resolve_ips(macs_list)
            hostnames = arp_table.resolve_hostnames(macs_list)
            row.append(", ".join(ips))
            row.append(", ".join(h for h in hostnames if h))
        row.extend([
            ", ".join(n.protocol for n in neighbors),
            ", ".join(n.device_id for n in neighbors),
            ", ".join(n.management_ip for n in neighbors),
            ", ".join(n.remote_port for n in neighbors),
            ", ".join(n.platform for n in neighbors),
        ])
        row.append(intf.description)
        ws.append(row)
        row_idx = ws.max_row
        fill = _link_fill(intf.link)
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(size=10)

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A5"
    return ws.title


def _write_lags_sheet(wb: Workbook, data: ParsedSwitchData) -> str:
    """Write a dedicated LAG summary for a switch."""
    base_name = f"{data.hostname or data.host} LAGs"[:31]
    ws = wb.create_sheet(title=base_name)
    _add_summary_link(ws)
    headers = [
        "LAG ID", "Name", "Mode", "Status", "Interface", "Trunk Type",
        "LACP Key", "Members", "Members Up", "Untagged VLAN", "Tagged VLANs",
    ]
    widths = [10, 28, 12, 12, 12, 16, 12, 42, 14, 22, 60]
    ws.append(headers)
    for col, header in enumerate(headers, start=1):
        _header_style(ws.cell(row=2, column=col), header)

    for lag in sorted(data.lags, key=lambda item: int(item.lag_id)):
        up_count = sum(state.lower() == "up" for state in lag.member_states.values())
        known_count = len(lag.member_states)
        member_health = f"{up_count}/{known_count}" if known_count else "Unknown"
        ws.append([
            lag.lag_id,
            lag.name,
            lag.mode,
            "Deployed" if lag.deployed else "Configured",
            lag.interface,
            lag.trunk_type,
            lag.lacp_key,
            ", ".join(lag.members),
            member_health,
            lag.untagged_vlan,
            lag.tagged_vlans,
        ])
        for cell in ws[ws.max_row]:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"
    return ws.title


def _write_raw_outputs_sheet(wb: Workbook, data: ParsedSwitchData) -> str:
    """Write command output one line per row, avoiding Excel cell limits."""
    base_name = f"{data.hostname or data.host} Raw"[:31]
    ws = wb.create_sheet(title=base_name)
    _add_summary_link(ws)
    ws.merge_cells("A2:C2")
    warning = ws["A2"]
    warning.value = (
        "SENSITIVE DATA — raw switch output may contain credentials, password "
        "hashes, SNMP communities, addresses, and configuration details."
    )
    warning.font = Font(bold=True, color="9C6500", size=10)
    warning.fill = PatternFill("solid", fgColor=COLOUR_WARNING)
    warning.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34

    headers = ["Command", "Line", "Raw output"]
    ws.append([])
    ws.append(headers)
    for col, header in enumerate(headers, start=1):
        _header_style(ws.cell(row=4, column=col), header)

    for command, output in data.raw_outputs.items():
        lines = output.splitlines() or [""]
        for line_number, line in enumerate(lines, start=1):
            ws.append([command, line_number, line])
            row = ws.max_row
            for cell in ws[row]:
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top")
                cell.font = Font(name="Courier New" if cell.column == 3 else "Arial", size=9)
            # Force raw device text to remain text even if it begins with '='.
            ws.cell(row=row, column=3).data_type = "s"

    _set_col_widths(ws, [26, 9, 120])
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:C{ws.max_row}"
    return ws.title


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary_sheet(
    wb: Workbook,
    all_data: list[ParsedSwitchData],
    sheet_links: list[dict[str, str | None]],
) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    headers = [
        "Switch", "Host", "Model", "Firmware", "Total Ports", "Up", "Down",
        "Disabled", "Main", "RAW", "LAG",
    ]
    widths = [20, 18, 24, 16, 12, 8, 8, 10, 10, 10, 10]

    ws.append(headers)
    for col, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 18

    for data, links in zip(all_data, sheet_links):
        total    = len(data.interfaces)
        up       = sum(1 for i in data.interfaces if i.link.lower() == "up")
        disabled = sum(1 for i in data.interfaces if i.link.lower() == "disabled")
        down     = total - up - disabled
        ws.append([
            data.hostname or data.host,
            data.host,
            data.model,
            data.firmware,
            total,
            up,
            down,
            disabled,
            "",
            "",
            "",
        ])
        row_idx = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(size=10)
        _add_sheet_link(ws.cell(row=row_idx, column=9), links["main"])
        _add_sheet_link(ws.cell(row=row_idx, column=10), links["raw"])
        _add_sheet_link(ws.cell(row=row_idx, column=11), links["lag"])

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export(
    all_data: list[ParsedSwitchData],
    output_path: Path,
    arp_table: ArpTable | None = None,
    include_raw_outputs: bool = False,
) -> Path:
    """Write all parsed switch data to an Excel workbook and return the path.

    If `arp_table` is provided, each switch tab gains IP and Hostname columns
    that map MACs from the switch's MAC table to entries in the ARP list.
    Raw CLI sheets are only written when explicitly requested."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    sheet_links: list[dict[str, str | None]] = []
    for data in all_data:
        main_sheet = _write_interfaces_sheet(wb, data, arp_table=arp_table)
        lag_sheet = None
        raw_sheet = None
        if data.lags:
            lag_sheet = _write_lags_sheet(wb, data)
        if include_raw_outputs and data.raw_outputs:
            raw_sheet = _write_raw_outputs_sheet(wb, data)
        sheet_links.append({"main": main_sheet, "raw": raw_sheet, "lag": lag_sheet})

    _write_summary_sheet(wb, all_data, sheet_links)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
