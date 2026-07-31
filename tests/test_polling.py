import queue
import threading
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock, patch

from netmiko.exceptions import ReadTimeout
from openpyxl import load_workbook

from pint_live.collectors import ruckus
from pint_live.core.polling import PollCancelled
from pint_live.models import InterfaceEntry, ParsedSwitchData
from pint_live.exporters.excel import export
from pint_live.parsers import ruckus as ruckus_parser
from pint_live.ui.app import PintLiveApp


def _parsed(host: str) -> ParsedSwitchData:
    return ParsedSwitchData(
        host=host,
        interfaces=[InterfaceEntry("1/1/1", "Up", "Forward", "Full", "1G", "No", "1")],
    )


def _job(host: str, collector, parser) -> dict:
    return {
        "host": host,
        "vendor": "Ruckus",
        "device_type": "ruckus_fastiron",
        "collector": collector,
        "parser": parser,
        "credentials": SimpleNamespace(username="user", password="pass"),
    }


def _worker_app() -> PintLiveApp:
    app = object.__new__(PintLiveApp)
    app._msg_queue = queue.Queue()
    app._stop_event = threading.Event()
    return app


def _done_message(app: PintLiveApp) -> tuple:
    messages = []
    while not app._msg_queue.empty():
        messages.append(app._msg_queue.get_nowait())
    return next(message for message in messages if message[0] == "done")


class PollWorkerTests(unittest.TestCase):
    def test_read_timeout_retries_with_fresh_connection_and_recovers(self):
        app = _worker_app()
        first_connection = Mock()
        second_connection = Mock()
        collector = Mock()
        collector.collect.side_effect = [ReadTimeout("prompt lost"), object()]
        parser = Mock()
        parser.parse.return_value = _parsed("10.0.0.1")

        with patch(
            "pint_live.ui.app.open_session",
            side_effect=[first_connection, second_connection],
        ) as connect:
            app._poll_worker([_job("10.0.0.1", collector, parser)])

        done = _done_message(app)
        self.assertEqual(len(done[1]), 1)
        self.assertEqual(done[2], [])
        self.assertFalse(done[3])
        self.assertEqual(connect.call_count, 2)
        first_connection.disconnect.assert_called_once()
        second_connection.disconnect.assert_called_once()

    def test_one_switch_failure_does_not_stop_remaining_switches(self):
        app = _worker_app()
        failed_connection = Mock()
        good_connection = Mock()
        failed_collector = Mock()
        failed_collector.collect.side_effect = RuntimeError("VPN dropped")
        good_collector = Mock()
        good_collector.collect.return_value = object()
        parser = Mock()
        parser.parse.return_value = _parsed("10.0.0.2")

        jobs = [
            _job("10.0.0.1", failed_collector, parser),
            _job("10.0.0.2", good_collector, parser),
        ]
        with patch(
            "pint_live.ui.app.open_session",
            side_effect=[failed_connection, good_connection],
        ):
            app._poll_worker(jobs)

        done = _done_message(app)
        self.assertEqual([result.host for result in done[1]], ["10.0.0.2"])
        self.assertEqual(done[2], [("10.0.0.1", "VPN dropped")])
        self.assertFalse(done[3])
        failed_connection.disconnect.assert_called_once()
        good_connection.disconnect.assert_called_once()

    def test_stop_before_next_switch_finishes_cleanly(self):
        app = _worker_app()
        app._stop_event.set()

        with patch("pint_live.ui.app.open_session") as connect:
            app._poll_worker([_job("10.0.0.1", Mock(), Mock())])

        done = _done_message(app)
        self.assertEqual(done[1], [])
        self.assertEqual(done[2], [])
        self.assertTrue(done[3])
        connect.assert_not_called()

    def test_empty_parser_result_is_reported_and_chain_continues(self):
        app = _worker_app()
        connections = [Mock(), Mock()]
        collector = Mock()
        collector.collect.return_value = object()
        parser = Mock()
        parser.parse.side_effect = [ParsedSwitchData(host="10.0.0.1"), _parsed("10.0.0.2")]

        with patch("pint_live.ui.app.open_session", side_effect=connections):
            app._poll_worker([
                _job("10.0.0.1", collector, parser),
                _job("10.0.0.2", collector, parser),
            ])

        done = _done_message(app)
        self.assertEqual([result.host for result in done[1]], ["10.0.0.2"])
        self.assertIn("No interfaces were recognised", done[2][0][1])


class CollectorCancellationTests(unittest.TestCase):
    def test_collector_honours_stop_between_commands(self):
        connection = Mock()
        with self.assertRaises(PollCancelled):
            ruckus.collect(connection, "10.0.0.1", lambda: True)
        connection.send_command.assert_not_called()

    def test_ruckus_commands_use_timing_reader(self):
        connection = Mock()
        connection.send_command_timing.side_effect = [
            "SW: Version 09.0.10h",
            "interfaces",
            "macs",
            "lags",
            "running config",
        ]

        raw = ruckus.collect(connection, "10.0.0.1")

        self.assertEqual(raw.version_output, "SW: Version 09.0.10h")
        self.assertEqual(connection.send_command_timing.call_count, 5)
        connection.send_command.assert_not_called()
        for call in connection.send_command_timing.call_args_list:
            self.assertEqual(call.kwargs, {"last_read": 3.0, "read_timeout": 120})


class RuckusParserTests(unittest.TestCase):
    def test_lags_include_members_health_and_vlan_assignments(self):
        raw = SimpleNamespace(
            host="10.0.0.1",
            version_output="System Name: CORE-01",
            interfaces_output="",
            mac_table_output="",
            lag_output='''
=== LAG "ACCESS-SW-01" ID 6 (dynamic Deployed) ===
LAG Configuration:
   Ports:         e 1/1/20 to 1/1/21 e 2/1/20 to 2/1/21
   Lag Interface: lg6
   Trunk Type:    hash-based
   LACP Key:      20006
Port       Link    State   Dupl Speed Trunk Tag Pvid Pri MAC Name
1/1/20     Up      Forward Full 25G   6     Yes 1    0   0000.0000.0000
1/1/21     Up      Forward Full 25G   6     Yes 1    0   0000.0000.0000
2/1/20     Down    None    None None  6     Yes 1    0   0000.0000.0000
2/1/21     Down    None    None None  6     Yes 1    0   0000.0000.0000
Port       [Sys P] [Port P] [ Key ] [Act]
''',
            running_config_output='''
lag ACCESS-SW-01 dynamic id 6
 ports ethe 1/1/20 to 1/1/21 ethe 2/1/20 to 2/1/21
!
vlan 1 name DEFAULT-VLAN by port
 untagged lag 6
!
vlan 36 name Dante by port
 tagged lag 6 to 7
!
''',
        )

        parsed = ruckus_parser.parse(raw)

        self.assertEqual(len(parsed.lags), 1)
        lag = parsed.lags[0]
        self.assertEqual(lag.lag_id, "6")
        self.assertEqual(lag.name, "ACCESS-SW-01")
        self.assertEqual(lag.members, ["1/1/20", "1/1/21", "2/1/20", "2/1/21"])
        self.assertEqual(lag.member_states["1/1/20"], "Up")
        self.assertEqual(lag.member_states["2/1/21"], "Down")
        self.assertEqual(lag.untagged_vlan, "1 (DEFAULT-VLAN)")
        self.assertEqual(lag.tagged_vlans, "36 (Dante)")

    def test_export_adds_lag_worksheet(self):
        raw = SimpleNamespace(
            host="10.0.0.1",
            version_output="System Name: CORE-01",
            interfaces_output="",
            mac_table_output="",
            lag_output='=== LAG "ACCESS-SW-01" ID 6 (dynamic Deployed) ===\n   Ports: e 1/1/20 e 2/1/20\n',
            running_config_output="vlan 36 name Dante by port\n tagged lag 6\n!\n",
        )
        parsed = ruckus_parser.parse(raw)

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "lags.xlsx"
            export([parsed], output)
            workbook = load_workbook(output, read_only=False)
            self.assertIn("CORE-01 LAGs", workbook.sheetnames)
            self.assertNotIn("CORE-01 Raw", workbook.sheetnames)
            lag_sheet = workbook["CORE-01 LAGs"]
            self.assertEqual(lag_sheet["A1"].value, "← Back to Summary")
            self.assertEqual(lag_sheet["A1"].hyperlink.target, "#'Summary'!A1")
            self.assertEqual(lag_sheet["A3"].value, "6")
            self.assertEqual(lag_sheet["K3"].value, "36 (Dante)")
            summary = workbook["Summary"]
            self.assertEqual(summary["I2"].hyperlink.target, "#'CORE-01'!A1")
            self.assertIsNone(summary["J2"].hyperlink)
            self.assertEqual(summary["K2"].hyperlink.target, "#'CORE-01 LAGs'!A1")
            workbook.close()

    def test_raw_output_worksheet_is_opt_in_and_line_based(self):
        raw = SimpleNamespace(
            host="10.0.0.1",
            version_output="System Name: CORE-01\n=not-a-formula",
            interfaces_output="",
            mac_table_output="",
            lag_output="",
            running_config_output="hostname CORE-01\n!\n",
        )
        parsed = ruckus_parser.parse(raw)

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "raw.xlsx"
            export([parsed], output, include_raw_outputs=True)
            workbook = load_workbook(output, read_only=False, data_only=False)
            self.assertIn("CORE-01 Raw", workbook.sheetnames)
            sheet = workbook["CORE-01 Raw"]
            self.assertEqual(sheet["A1"].hyperlink.target, "#'Summary'!A1")
            self.assertEqual(sheet["A5"].value, "show version")
            self.assertEqual(sheet["B5"].value, 1)
            self.assertEqual(sheet["C6"].value, "=not-a-formula")
            self.assertEqual(sheet["C6"].data_type, "s")
            self.assertEqual(
                workbook["Summary"]["J2"].hyperlink.target,
                "#'CORE-01 Raw'!A1",
            )
            workbook.close()

    def test_hostname_falls_back_to_running_config(self):
        raw = SimpleNamespace(
            host="10.0.0.1",
            version_output="SW: Version 09.0.10",
            interfaces_output="",
            mac_table_output="",
            running_config_output="hostname LAB-ICX-01\n!\n",
        )

        parsed = ruckus_parser.parse(raw)

        self.assertEqual(parsed.hostname, "LAB-ICX-01")

    def test_show_version_hostname_takes_precedence(self):
        raw = SimpleNamespace(
            host="10.0.0.1",
            version_output="System Name: LIVE-NAME",
            interfaces_output="",
            mac_table_output="",
            running_config_output="hostname CONFIG-NAME\n!\n",
        )

        parsed = ruckus_parser.parse(raw)

        self.assertEqual(parsed.hostname, "LIVE-NAME")


if __name__ == "__main__":
    unittest.main()
